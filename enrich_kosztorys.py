import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy

SRC = '/Users/michal/workspace/ai-for-business/Kosztorys.xlsx'
DST = '/Users/michal/workspace/ai-for-business/Kosztorys.xlsx'

wb = openpyxl.load_workbook(SRC)

# ─── helpers ──────────────────────────────────────────────────────────────────
def font(size=11, bold=False, name='Aptos'):
    return Font(name=name, size=size, bold=bold)

hdr_font  = font(size=11, bold=True)
sec_font  = font(size=13, bold=True, name='Calibri')
title_font = font(size=14, bold=True, name='Calibri')
body_font  = font(size=11, name='Aptos')
body_sm    = font(size=10, name='Aptos')

wrap = Alignment(wrap_text=True, vertical='top')
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_row(ws, row, col_start, values, fonts=None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=v)
        if fonts and i < len(fonts) and fonts[i]:
            c.font = fonts[i]
        c.alignment = wrap

def merge_and_set(ws, r1, c1, r2, c2, value, fnt=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    if fnt:
        c.font = fnt
    c.alignment = wrap

# ─── fetch totals from Podsumowanie ──────────────────────────────────────────
ps = wb['Podsumowanie']
totals = {}
# TOTAL.01 – row 5..11, TOTAL.02 – row 13..19, TOTAL.03 – row 21..27
# TOTAL.04 – row 29..35, TOTAL.05 – row 37..43, TOTAL.06 – row 45..51
total_rows = [
    ('TOTAL.01', 'Bankowość Centralna',         5, 11),
    ('TOTAL.02', 'Capital Markets AI SDLC',     13, 19),
    ('TOTAL.03', 'Capital Markets Refactor',     21, 27),
    ('TOTAL.04', 'Pakiet szkoleń',              29, 35),
    ('TOTAL.05', 'Gwarancja / Bufor',           37, 43),
    ('TOTAL.06', 'Narzędzia i licencje',        45, 51),
]
for code, label, row_start, row_sum in total_rows:
    cell = ps.cell(row=row_sum, column=5)  # column E
    totals[code] = {'label': label, 'value': cell.value, 'row_sum': row_sum}

grand_total = ps.cell(row=56, column=5).value

# ─── 1. Offer Overview sheet ─────────────────────────────────────────────────
ov = wb.create_sheet('Offer Overview', 0)

r = 1
merge_and_set(ov, r, 1, r, 6, 'AI SDLC — Oferta wdrożenia', title_font)
r = 3
merge_and_set(ov, r, 1, r, 6, 'Executive Summary', sec_font)
r = 4
merge_and_set(ov, r, 1, r+1, 6,
    'Bottega proponuje kompleksowe wdrożenie inżynierii oprogramowania '
    'wspomaganej sztuczną inteligencją (AI SDLC) w dwóch obszarach '
    'biznesowych: Bankowość Centralna oraz Capital Markets, wraz z '
    'opcjonalną technologiczną restrukturyzacją wybranych modułów.\n\n'
    'Oferta obejmuje budowę baz wiedzy i dokumentacji zoptymalizowanej pod '
    'AI, konfigurację narzędzi i instrukcji dla analityków, developerów '
    'i reviewerów, wdrożenie automatyzacji CI/CD oraz security gates, '
    'a także szkolenia i transfer know-how.\n\n'
    f'Łączna wartość inwestycji: {grand_total} PLN netto.',
    body_font)

r = 7
merge_and_set(ov, r, 1, r, 6, 'Podejście i fazy wdrożenia', sec_font)
r = 8
merge_and_set(ov, r, 1, r, 6,
    'Nasze podejście to stopniowa ewolucja, a nie rewolucja. Integrujemy '
    'AI tam, gdzie przynosi wymierny zwrot z inwestycji i odciąża '
    'inżynierów od powtarzalnych zadań, unikając pułapek związanych '
    'z bezpieczeństwem, halucynacjami modeli i obniżeniem jakości kodu.',
    body_font)
r = 9; ov.cell(row=r, column=1, value='Faza 1 — Audit i strategia').font = hdr_font
ov.cell(row=r, column=2).value = 'Analiza procesu wytwórczego, stosu technologicznego i kultury inżynierskiej. Identyfikacja obszarów o najwyższym potencjale zwrotu. Wynik: roadmapa wdrożenia.'
ov.cell(row=r, column=2).font = body_font; ov.cell(row=r, column=2).alignment = wrap
r = 10; ov.cell(row=r, column=1, value='Faza 2 — Proof of Value').font = hdr_font
ov.cell(row=r, column=2).value = 'Pilotaż na wybranym zespole lub module. Praca własna Bottega (przygotowanie narzędzi, dokumentacji) + pair programming online (transfer know-how). Wynik: działające rozwiązanie + udowodniona wartość.'
ov.cell(row=r, column=2).font = body_font; ov.cell(row=r, column=2).alignment = wrap
r = 11; ov.cell(row=r, column=1, value='Faza 3 — Skalowanie kompetencji').font = hdr_font
ov.cell(row=r, column=2).value = 'Rozszerzenie praktyk na całą organizację. Przekazanie licencji z bezterminowym prawem do użytku. Szkolenia dla analityków, PM i developerów.'
ov.cell(row=r, column=2).font = body_font; ov.cell(row=r, column=2).alignment = wrap

r = 13
merge_and_set(ov, r, 1, r, 6, 'Zakres wdrożenia', sec_font)
scope_descs = [
    ('TOTAL.01', 'Bankowość Centralna — AI SDLC',
     'Pełen zestaw praktyk AI SDLC: skonsolidowanie dokumentacji funkcji '
     'systemu (IK.1), mapowanie funkcjonalności na kod i architekturę '
     '(IK.2), dokumentacja frontendu i backendu dla AI (IK.4, IK.5), '
     'instrukcje AI do analizy regulatorów KNF/UE (EK.1), narzędzia dla '
     'analityków i developerów (TL.1–TL.5).\n'
     'Efekt: wspólny język z AI, szybsza analiza impaktu, eliminacja '
     'błędów z nieaktualnej dokumentacji.'),
    ('TOTAL.02', 'Capital Markets — AI SDLC',
     'Analogiczny zakres jak TOTAL.01, dostosowany do specyfiki '
     'Capital Markets (regulator GPW). Większy wymiar pracy własnej '
     'Bottega by odciążyć zespół klienta.'),
    ('TOTAL.03', 'Capital Markets — Refaktoryzacja',
     'Standaryzacja CI/CD (Jenkinsfile, shared library), migracja '
     'Ant→Maven/Gradle, automatyzacja deploy QA, ujednolicenie '
     'artefaktów, iOS build/release (Fastlane), security gates '
     '(Fortify, Dependency-Track), plan migracji z wyceną etapową.'),
    ('TOTAL.04', 'Pakiet szkoleń',
     'Szkolenie AI SDLC dla analityków i PM (2 dni) oraz dla '
     'developerów (3 dni). Bazuje na realnych zadaniach i kodzie '
     'projektów.'),
    ('TOTAL.05', 'Gwarancja 1 rok / Bufor',
     'Pakiet dodatkowych godzin na bieżące wsparcie, korekty '
     'i niespodziewane wyzwania po wdrożeniu.'),
    ('TOTAL.06', 'Narzędzia i licencje',
     'Przekazanie licencji na narzędzia AI SDLC z bezterminowymi '
     'prawami do wewnętrznego użytku i własnego rozwoju.'),
]
for code, label, desc in scope_descs:
    r += 1
    ov.cell(row=r, column=1, value=f'{code} — {label}').font = hdr_font
    ov.cell(row=r, column=2, value=desc).font = body_font
    ov.cell(row=r, column=2).alignment = wrap
    amt = totals.get(code, {}).get('value')
    if amt is not None:
        ov.cell(row=r, column=5, value=amt).font = body_font

r += 2
merge_and_set(ov, r, 1, r, 6, 'Efekty biznesowe', sec_font)
outcomes = [
    ('Przyspieszenie Time-to-Market',
     'Automatyzacja powtarzalnych zadań skraca czas dostarczania '
     'funkcjonalności. AI przejmuje żmudne czynności, inżynierowie '
     'skupiają się na pracy twórczej.\n→ TOTAL.01, TOTAL.02, TOTAL.03'),
    ('Redukcja długu technicznego',
     'Systematyczna poprawa jakości kodu i architektury przez AI '
     'Review (bugs, architektura, complexicity, zgodność ze specyfikacją).\n'
     '→ TOTAL.03, TOTAL.01, TOTAL.02'),
    ('Retencja wiedzy projektowej',
     'AI jako "kustosz wiedzy" — skonsolidowana dokumentacja, baza '
     'wiedzy mapująca funkcjonalność na kod. Szybsze wdrażanie nowych '
     'pracowników, minimalizacja ryzyka utraty wiedzy przy rotacji.\n'
     '→ TOTAL.01, TOTAL.02, TOTAL.06'),
    ('Skalowanie kompetencji',
     'Ujednolicenie standardów pracy. Dzięki narzędziom i instrukcjom '
     'AI każdy inżynier pracuje według tych samych wzorców.\n'
     '→ TOTAL.04, TOTAL.01, TOTAL.02'),
    ('Bezpieczeństwo danych',
     'Pełna kontrola nad kodem źródłowym i dokumentacją. Gating '
     'bezpieczeństwa w pipeline. Opcje: Cloud Native, Private Cloud '
     '(VPC) lub On-Premise (Air-Gapped).\n→ TOTAL.03, TOTAL.06'),
]
for label, desc in outcomes:
    r += 1
    ov.cell(row=r, column=1, value=label).font = hdr_font
    ov.cell(row=r, column=2, value=desc).font = body_font
    ov.cell(row=r, column=2).alignment = wrap

r += 2
merge_and_set(ov, r, 1, r, 6, 'Model dostawy', sec_font)
r += 1
ov.cell(row=r, column=1, value='Formy współpracy').font = hdr_font
r += 1
delivery_rows = [
    ('Praca własna Bottega', 'Eksperci Bottega realizują zadania samodzielnie',
     'Budowa narzędzi, dokumentacji, instrukcji'),
    ('Pair programming online', 'Wspólna praca eksperta Bottega z inżynierem klienta',
     'Transfer kompetencji: context engineering, refactoring, pipeline'),
    ('Review po stronie klienta', 'Klient weryfikuje efekty przed zamknięciem',
     'Każda dostarczona pozycja'),
]
for form, desc, when in delivery_rows:
    ov.cell(row=r, column=1, value=form).font = body_sm
    ov.cell(row=r, column=2, value=desc).font = body_sm
    ov.cell(row=r, column=4, value=when).font = body_sm
    r += 1
r += 1
ov.cell(row=r, column=1, value='Środowiska docelowe').font = hdr_font
r += 1
envs = [
    'Cloud Native (Enterprise API) — dla firm akceptujących chmurę z BAA',
    'Private Cloud (VPC) — dla organizacji z regulacjami korporacyjnymi',
    'On-Premise (Air-Gapped) — dla banków i sektora publicznego',
]
for e in envs:
    ov.cell(row=r, column=1, value=e).font = body_font
    r += 1

r += 1
merge_and_set(ov, r, 1, r, 6, 'Ryzyka i mitigacje', sec_font)
r += 1
risks = [
    ('Halucynacje modeli AI',
     'Context engineering — budowa precyzyjnego kontekstu dla LLM; '
     'instrukcje AI wymuszające weryfikację faktów; human-in-the-loop'),
    ('Obniżenie jakości kodu',
     'AI Review w pipeline (bugs, architektura, complexicity, '
     'zgodność ze specyfikacją); gating security; code review'),
    ('Wyciek danych / kodu źródłowego',
     'Opcje On-Premise (Air-Gapped) lub Private Cloud (VPC); '
     'polityka danych i instrukcje AI wymuszające lokalne przetwarzanie'),
    ('Odrzucenie przez zespół',
     'Stopniowe wprowadzanie (ewolucja, nie rewolucja); wartość na pilocie; '
     'szkolenia i mentoring; odciążenie od powtarzalnych zadań'),
    ('Ponoszenie zbędnych kosztów',
     'Harness engineering — wyższa jakość przy mniejszych modelach; '
     'wycinanie zbędnego outputu i logów; szkolenia z efektywnego '
     'korzystania z AI (TOTAL.04)'),
    ('Uzależnienie od dostawcy',
     'Przekazanie narzędzi i licencji na własność; bezterminowe prawa; '
     'transfer know-how przez pair programming'),
]
for risk, mitig in risks:
    ov.cell(row=r, column=1, value=risk).font = hdr_font
    ov.cell(row=r, column=2, value=mitig).font = body_font
    ov.cell(row=r, column=2).alignment = wrap
    r += 1

r += 1
merge_and_set(ov, r, 1, r, 6, 'Dlaczego Bottega', sec_font)
r += 1
bottega = [
    'Doświadczenie w bankowości konsumenckiej i inwestycyjnej — projekty '
    'transformacyjne w sektorze finansowym; znajomość regulacji KNF, GPW i UE',
    '15 lat doświadczenia w szkoleniach i doradztwie dla zespołów deweloperskich',
    '78 aktywnych trenerów o modelu kompetencji Π (Pi) — silny filar techniczny '
    'oraz umiejętności miękkie, talent edukacyjny, myślenie strategiczne',
    'Ponad 12 000 przeszkolonych uczestników, 150 dni szkoleń rocznie',
    '43 klientów z branż: finansowa, medyczna, ubezpieczeniowa, energetyczna, '
    'zbrojeniowa, e-commerce',
    'Eksperci: Sławomir Sobótka, Jakub Nabrdalik, Łukasz Szydło, Mariusz Gil',
]
for b in bottega:
    ov.cell(row=r, column=1, value=b).font = body_font
    r += 1
r += 1
merge_and_set(ov, r, 1, r, 4,
    'Wybrani klienci: Intel Polska • Orange Polska • Roche Polska • '
    'Asseco Polska • Tomtom Polska • Capgemini • Comarch • Luxoft • '
    'Motorola • PZU • Tieto • Atos', body_sm)

# Set column widths for Offer Overview
ov.column_dimensions['A'].width = 32
ov.column_dimensions['B'].width = 55
ov.column_dimensions['C'].width = 12
ov.column_dimensions['D'].width = 12
ov.column_dimensions['E'].width = 18
ov.column_dimensions['F'].width = 12

# ─── 2. Opis oferty column in Podsumowanie ───────────────────────────────────
PS_DESC = {
    5:    ('TOTAL.01 — Bankowość Centralna: AI SDLC + Know-How',
           'Pełen zestaw praktyk AI SDLC: skonsolidowanie dokumentacji '
           'funkcji systemu (IK.1), mapowanie funkcjonalności na kod '
           'i architekturę (IK.2), dokumentacja frontendu i backendu '
           'zoptymalizowana pod AI (IK.4, IK.5), instrukcje AI do analizy '
           'regulatorów KNF/UE (EK.1) oraz narzędzia i instrukcje dla '
           'analityków i developerów (TL.1–TL.5). Efekt: analitycy '
           'i developerzy zyskują wspólny język z AI.'),
    6:    None,  # continuation of TOTAL.01 description
    13:   ('TOTAL.02 — Capital Markets: AI SDLC + Know-How',
           'Analogiczny zakres jak dla Bankowości Centralnej, dostosowany '
           'do specyfiki BU Capital Markets (w tym regulator GPW). '
           'Większy wymiar pracy własnej Bottega by odciążyć zespół klienta.'),
    14:   None,
    21:   ('TOTAL.03 — Capital Markets: Technologiczna restrukturyzacja',
           'Standaryzacja CI/CD na Jenkinsfile (REF.1), migracja builda '
           'Ant→Maven/Gradle (REF.2), automatyzacja deploy QA (REF.3), '
           'ujednolicenie artefaktów (REF.4), iOS build/release przez '
           'Fastlane (REF.5), Fortify/Dependency-Track z gatingiem (REF.6), '
           'plan migracji z wyceną etapową (REF.RE.1).'),
    22:   None,
    29:   ('TOTAL.04 — Pakiet szkoleń',
           'Szkolenia dla szerszego grona pracowników: AI SDLC dla '
           'analityków i PM (2 dni) oraz dla developerów (3 dni). '
           'Bazują na realnych zadaniach i kodzie projektów.'),
    30:   None,
    37:   ('TOTAL.05 — Gwarancja 1 rok / Bufor',
           'Pakiet dodatkowych godzin na bieżące wsparcie, korekty '
           'i niespodziewane wyzwania po wdrożeniu.'),
    38:   None,
    45:   ('TOTAL.06 — Przekazanie narzędzi i licencji',
           'Licencje na narzędzia AI SDLC z bezterminowymi prawami '
           'do wewnętrznego użytku i własnego rozwoju.'),
    46:   None,
}
ps.cell(row=4, column=7, value='Opis oferty').font = hdr_font

for row_num, val in PS_DESC.items():
    if val is None:
        continue
    title, desc = val
    cell = ps.cell(row=row_num, column=7, value=f'{title}\n{desc}')
    cell.font = body_sm
    cell.alignment = wrap

ps.column_dimensions['G'].width = 55

# ─── 3. Value proposition column in detailed sheets ─────────────────────────
VP = {
    'Central Banking': {
        'IK.1': 'Analityk: pewność że AI rozumie aktualny stan systemu. '
                'Developer: AI i agenci wiedzą co jest już zaimplementowane. '
                'Reviewer: kontekst do oceny impaktu. QA: pełny kontekst systemu.',
        'IK.2': 'Analityk: AI wskazuje miejsca zmian po opisie. '
                'Developer: wie dokładnie gdzie zmieniać kod. '
                'Reviewer: ocenia pełen impact. QA: wie co testować.',
        'IK.3': 'Liderzy: estymaty oparte na wartościach eksperckich.',
        'IK.4': 'Analityk: szybsza ocena impaktu na frontend. '
                'Developer: AI generuje kod zgodny z architekturą. '
                'Reviewer: zgodność z zasadami. QA: spójne testy UI.',
        'IK.5': 'Analityk: szybsza ocena impaktu na backend. '
                'Developer: AI generuje kod zgodny z wzorcami backendu. '
                'Reviewer: zgodność architektoniczna. QA: testy zgodne ze wzorcami.',
        'IK.7': 'Analityk: brakujące scenariusze z AI. '
                'Developer: scenariusze z danymi. '
                'Reviewer: sprawdza pokrycie. QA: gotowe e2e.',
        'EK.1': 'Analityk: AI rozbudowuje wymagania regulacyjne, generuje '
                'pytania doprecyzowujące, weryfikuje analizę vs źródła.',
        'TL.1': 'Analityk: dedykowane środowisko AI do analizy i review dokumentów.',
    },
    'Capital Markets': {
        'IK.1': 'Analityk: pewność że AI rozumie aktualny stan systemu. '
                'Developer: AI i agenci wiedzą co jest już zaimplementowane. '
                'Reviewer: kontekst do oceny impaktu. QA: pełny kontekst systemu.',
        'IK.2': 'Analityk: AI wskazuje miejsca zmian po opisie. '
                'Developer: wie dokładnie gdzie zmieniać kod. '
                'Reviewer: ocenia pełen impact. QA: wie co testować.',
        'IK.3': 'Liderzy: estymaty oparte na wartościach eksperckich.',
        'IK.4': 'Analityk: szybsza ocena impaktu na frontend. '
                'Developer: AI generuje kod zgodny z architekturą. '
                'Reviewer: zgodność z zasadami. QA: spójne testy UI.',
        'IK.5': 'Analityk: szybsza ocena impaktu na backend. '
                'Developer: AI generuje kod zgodny z wzorcami backendu. '
                'Reviewer: zgodność architektoniczna. QA: testy zgodne ze wzorcami.',
        'IK.7': 'Analityk: brakujące scenariusze z AI. '
                'Developer: scenariusze z danymi. '
                'Reviewer: sprawdza pokrycie. QA: gotowe e2e.',
        'EK.1': 'Analityk: AI rozbudowuje wymagania regulacyjne KNF/GPW, '
                'generuje pytania, weryfikuje analizę.',
        'TL.1': 'Analityk: dedykowane środowisko AI do analizy i review dokumentów.',
    },
    'Refactor Capital Markets': {
        'REF.RE.1': 'Klient: realny plan migracji z etapami i wyceną zamiast estymacji w ciemno.',
        'REF.1': 'Developer + AI: jeden znajomy flow CI/CD dla wszystkich repozytoriów.',
        'REF.2': 'Developer: szybsze buildy, dependency management, security scanning.',
        'REF.3': 'Developer: self-service deploy, rollback. Koniec wąskiego gardła.',
        'REF.4': 'Tech lead: w każdej chwili wiadomo co i kiedy trafiło na środowisko.',
        'REF.5': 'iOS developer: build i podpisywanie w pipeline. Koniec ręcznych obejść.',
        'REF.6': 'Security: stałe skany SAST/SCA jako warunki blokujące w pipeline.',
    },
    'Narzędzia': {
        'RE.1': 'Rekomendacje narzędzi AI dopasowane do istniejących subskrypcji i infrastruktury.',
        'IK.1': 'Narzędzie do konsolidacji dokumentacji funkcji systemu dla AI.',
        'IK.2': 'Narzędzie mapowania funkcjonalności na kod i architekturę.',
        'IK.3': 'Narzędzie heurystycznego szacowania pracochłonności.',
        'IK.4': 'Narzędzie dokumentacji architektury frontendu dla AI.',
        'IK.5': 'Narzędzie dokumentacji architektury backendu dla AI.',
        'IK.7': 'Narzędzie implementacji scenariuszy e2e.',
        'EK.1': 'Narzędzie AI do pracy ze źródłami regulatorów KNF/UE.',
        'TL.1': 'Środowisko AI dla analityków.',
        'TL.2': 'Instrukcje AI do gap analysis i spike analysis.',
        'TL.2.1': 'Instrukcje AI do review analizy vs KNF/UE.',
        'TL.3': 'Generowanie diagramów BPMN/UML z AI.',
        'TL.4': 'Generowanie UI/UX low fidelity z AI.',
        'TL.5': 'AI proponuje scenariusze testowe z analizy.',
        'TL.29': 'Integracja z Enterprise Architect.',
        'TL.30': 'Formatowanie specyfikacji dla AI.',
        'TL.6': 'Reverse engineering mapowania funkcja→kod.',
        'TL.9': 'AI odpyta bazę Oracle (readonly).',
        'TL.10': 'Reverse engineering zasad architektonicznych.',
        'TL.13': 'AI robi hot swap, redeploy, podgląda logi.',
        'TL.14': 'AI uruchamia pipeline Jenkins.',
        'TL.31': 'AI lokalny troubleshooting.',
        'TL.32': 'AI wywołuje remote beans dla smoke testów.',
        'TL.11': 'Automatyczne zbiorcze review w pipeline.',
        'TL.15': 'Przyrostowe wykrywanie potencjalnych bugów.',
        'TL.16': 'Przyrostowe sprawdzenie zgodności z architekturą.',
        'TL.17': 'Przyrostowa kontrola złożoności kodu.',
        'TL.18': 'Sprawdzenie czy kod realizuje specyfikację.',
    },
}

def find_code_row(ws, code):
    """Find row number where column A matches the given code."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == code:
            return r
    return None

for sheet_name, vp_map in VP.items():
    ws = wb[sheet_name]
    # Add header in column I (col 9)
    ws.cell(row=4, column=9, value='Value proposition').font = hdr_font
    for code_text, vp_text in vp_map.items():
        row = find_code_row(ws, code_text)
        if row:
            cell = ws.cell(row=row, column=9, value=vp_text)
            cell.font = body_sm
            cell.alignment = wrap
    ws.column_dimensions['I'].width = 50

# ─── save ────────────────────────────────────────────────────────────────────
wb.save(DST)
print('Done — enriched Kosztorys.xlsx saved.')
